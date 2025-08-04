# xlsx_parser.py
# Excel文档解析器（待实现）

import os
from typing import List, Dict, Optional, Tuple
import pandas as pd
from utils.logger import logger
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.zhipu_client import zhipu_complete_async, parse_json_response
from utils.chunk_prompts import (
    SYSTEM_PROMPTS,
    STRUCTURED_PROMPTS,
    STRUCTURED_PROMPTS_WITH_CONTEXT,
)
from utils.config import LLM_CONFIG
from .fragment_config import FragmentConfig, TableProcessingConfig
import asyncio


class XlsxFileParser:
    """Excel（.xlsx）文件解析器，输出分块结构，支持多Sheet和多表格。"""

    def __init__(self, fragment_config: Optional[FragmentConfig] = None):
        """初始化解析器"""
        self.table_config = fragment_config.table_processing if fragment_config else TableProcessingConfig()

    def parse(self, file_path: str) -> List[Dict]:
        """
        解析Excel文件，输出分块结构，每个分块包含type、content、metadata、context、parent_id。
        """
        logger.info(f"开始解析Excel文档: {file_path}")
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return []
        doc_id = os.path.basename(file_path)
        all_chunks: List[Dict] = []
        xls = None
        wb = None
        try:
            xls = pd.ExcelFile(file_path)
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                ws = wb[sheet_name]
                table_blocks = self._split_table_blocks(ws, df)
                merged_cells = self._get_merged_cells(ws)
                for block_idx, (start_row, end_row) in enumerate(table_blocks):
                    table_id = f"{sheet_name}_table_{block_idx+1}"
                    block_df = df.iloc[start_row : end_row + 1, :].copy()
                    block_df = self._drop_empty_cols(block_df)
                    if block_df.empty:
                        continue
                    header_rows = self._detect_header_rows(ws, start_row, end_row)
                    
                    # 传递worksheet信息给表头处理
                    headers = self._get_multi_headers_with_worksheet(block_df, header_rows, ws, start_row)
                    
                    # 根据配置获取表格格式
                    table_content, table_headers, table_merged_cells = self._convert_table_to_format(block_df, header_rows, ws, start_row)
                    
                    parent_table_info = self._get_parent_table_info(
                        headers, merged_cells
                    )
                    context = self._get_context_for_table(df, start_row, end_row)
                    
                    # 根据配置决定是否生成完整表格块
                    if self.table_config.table_chunking_strategy in ["full_only", "full_and_rows"]:
                        table_chunk = {
                            "type": "table_full",
                            "content": table_content,
                            "metadata": {
                                "doc_id": doc_id,
                                "sheet": sheet_name,
                                "table_id": table_id,
                                "start_row": int(start_row),
                                "end_row": int(end_row),
                                "header_rows": header_rows,
                                "header": headers,
                                "merged_cells": merged_cells,
                                "parent_table_info": parent_table_info,
                                "table_format": self.table_config.table_format,
                            },
                            "context": context,
                            "parent_id": None,
                        }
                        all_chunks.append(table_chunk)
                    
                    # 根据配置决定是否生成行级分块
                    if self.table_config.table_chunking_strategy == "full_and_rows":
                        row_chunks = self._generate_table_row_chunks(
                            block_df, doc_id, sheet_name, table_id, start_row, 
                            headers, parent_table_info, header_rows
                        )
                        all_chunks.extend(row_chunks)
        except Exception as e:
            logger.error(f"解析Excel文件出错: {file_path}, 错误: {str(e)}")
            return []
        finally:
            # 确保文件正确关闭
            if xls is not None:
                xls.close()
            if wb is not None:
                wb.close()

        # 为每个分块添加chunk_id
        for idx, chunk in enumerate(all_chunks):
            chunk["chunk_id"] = f"{doc_id}_{idx+1}"

        # 调用LLM增强所有分块
        try:
            import asyncio
            import nest_asyncio

            # 应用nest_asyncio以支持嵌套事件循环
            nest_asyncio.apply()

            # 检查是否已有运行的事件循环
            try:
                loop = asyncio.get_running_loop()
                # 如果已有事件循环，创建任务
                task = loop.create_task(enhance_all_chunks(all_chunks))
                enhanced_chunks = loop.run_until_complete(task)
            except RuntimeError:
                # 如果没有运行的事件循环，使用asyncio.run
                enhanced_chunks = asyncio.run(enhance_all_chunks(all_chunks))

            logger.info(f"成功增强 {len(enhanced_chunks)} 个分块")
            return enhanced_chunks
        except Exception as e:
            logger.error(f"LLM增强分块失败: {str(e)}")
            return all_chunks  # 如果增强失败，返回原始分块

    def _convert_table_to_format(self, df: pd.DataFrame, header_rows: int, ws: Worksheet = None, start_row: int = 0) -> Tuple[str, List[str], List[Dict]]:
        """根据配置将表格转换为指定格式，正确处理合并单元格"""
        try:
            if self.table_config.table_format == "markdown":
                return self._df_to_markdown_multiheader(df, header_rows, ws, start_row)
            else:
                return self._df_to_html_multiheader(df, header_rows, ws, start_row)
        except Exception as e:
            logger.warning(f"格式转换失败，回退到HTML格式: {str(e)}")
            return self._df_to_html_multiheader(df, header_rows, ws, start_row)

    def _generate_table_row_chunks(
        self, 
        block_df: pd.DataFrame, 
        doc_id: str, 
        sheet_name: str, 
        table_id: str, 
        start_row: int,
        headers: List[str], 
        parent_table_info: str,
        header_rows: int
    ) -> List[Dict]:
        """生成表格行分块"""
        row_chunks = []
        for r_idx in range(header_rows, block_df.shape[0]):
            row_content = self._row_to_format(block_df.iloc[r_idx], headers)
            row_context = self._get_context_for_row(
                block_df, r_idx, header_rows
            )
            row_chunk = {
                "type": "table_row",
                "content": row_content,
                "metadata": {
                    "doc_id": doc_id,
                    "sheet": sheet_name,
                    "row": int(start_row + r_idx + 1),
                    "table_id": table_id,
                    "header": headers,
                    "parent_table_info": parent_table_info,
                    "table_format": self.table_config.table_format,
                },
                "context": row_context,
                "parent_id": table_id,
            }
            row_chunks.append(row_chunk)
        return row_chunks

    def _row_to_format(self, row: pd.Series, headers: List[str]) -> str:
        """根据配置将表格行转换为指定格式"""
        if self.table_config.table_format == "markdown":
            return self._row_to_markdown(row, headers)
        else:
            return self._row_to_html(row, headers)

    def _detect_header_rows(
        self, ws: Worksheet, start_row: int, end_row: int, max_header_rows: int = 3
    ) -> int:
        """启发式分析[start_row, end_row]区间内的表头行数，基于合并单元格分布。"""
        merge_counts = [0] * (end_row - start_row + 1)
        for mc in ws.merged_cells.ranges:
            for row in range(
                max(mc.min_row, start_row + 1), min(mc.max_row, end_row + 1) + 1
            ):
                merge_counts[row - (start_row + 1)] += 1
        header_rows = 0
        for i in range(min(max_header_rows, len(merge_counts))):
            if merge_counts[i] > 0:
                header_rows += 1
            else:
                break
        return header_rows if header_rows > 0 else 1

    def _split_table_blocks(
        self, ws: Worksheet, df: pd.DataFrame
    ) -> List[Tuple[int, int]]:
        """结合合并单元格、空行、非空密度智能分块。返回每个表格的起止行索引。"""
        blocks: List[Tuple[int, int]] = []
        in_block = False
        start = 0
        for idx, row in df.iterrows():
            has_data = row.notna().any()
            has_merge = self._row_has_merge(ws, idx + 1)
            if has_data or has_merge:
                if not in_block:
                    start = idx
                    in_block = True
            else:
                if in_block:
                    blocks.append((start, idx - 1))
                    in_block = False
        if in_block:
            blocks.append((start, df.shape[0] - 1))
        return blocks

    def _row_has_merge(self, ws: Worksheet, row_idx: int) -> bool:
        """判断某一行是否有合并单元格。"""
        for mc in ws.merged_cells.ranges:
            if mc.min_row <= row_idx <= mc.max_row:
                return True
        return False

    def _drop_empty_cols(self, df: pd.DataFrame) -> pd.DataFrame:
        """去除全空列。"""
        return df.dropna(axis=1, how="all")

    def _get_multi_headers(self, df: pd.DataFrame, header_rows: int) -> List[str]:
        """获取多级表头，正确处理合并单元格的层次结构。"""
        if df.shape[0] < header_rows:
            return []
        
        # 简化实现：直接拼接表头行的内容
        headers = []
        for col in range(df.shape[1]):
            header_parts = []
            for row in range(header_rows):
                cell_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                if cell_value.strip():
                    header_parts.append(cell_value)
            header = "/".join(header_parts) if header_parts else ""
            headers.append(header)
        
        return headers

    def _get_multi_headers_with_worksheet(self, df: pd.DataFrame, header_rows: int, ws: Worksheet, start_row: int) -> List[str]:
        """获取多级表头，并传递worksheet信息。"""
        if df.shape[0] < header_rows:
            return []
        
        # 获取合并单元格信息（从Excel文件获取）
        merged_cells = self._get_excel_merged_cells_with_worksheet(ws, start_row, header_rows)
        
        # 构建表头映射
        header_mapping = self._build_header_mapping_with_worksheet(df, header_rows, merged_cells, ws, start_row)
        
        # 生成最终表头
        headers = []
        for col in range(df.shape[1]):
            header = header_mapping.get(col, "")
            headers.append(header)
        
        return headers

    def _get_excel_merged_cells_with_worksheet(self, ws: Worksheet, start_row: int, header_rows: int) -> List[Dict]:
        """从Excel worksheet获取表头区域的合并单元格信息。"""
        merged_cells = []
        
        # 检查表头区域的合并单元格
        for row in range(start_row + 1, start_row + header_rows + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and self._is_merged_cell_header(str(cell.value)):
                    # 检查是否有合并单元格
                    colspan = self._get_cell_colspan(ws, row, col)
                    if colspan > 1:
                        merged_cells.append({
                            "row": row - start_row - 1,  # 转换为DataFrame索引
                            "col": col - 1,  # 转换为DataFrame索引
                            "colspan": colspan,
                            "text": str(cell.value)
                        })
        
        return merged_cells

    def _get_cell_colspan(self, ws: Worksheet, row: int, col: int) -> int:
        """获取单元格的列跨度。"""
        # 检查该单元格是否被合并
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return merged_range.max_col - merged_range.min_col + 1
        return 1

    def _build_header_mapping_with_worksheet(self, df: pd.DataFrame, header_rows: int, merged_cells: List[Dict], ws: Worksheet, start_row: int) -> Dict[int, str]:
        """构建列到表头的映射，处理合并单元格的层次结构。"""
        header_mapping = {}
        
        # 处理合并单元格
        for merge_info in merged_cells:
            parent_text = merge_info["text"]
            start_col = merge_info["col"]
            colspan = merge_info["colspan"]
            
            # 为合并单元格的子列分配表头
            for col_offset in range(colspan):
                col = start_col + col_offset
                if col < df.shape[1]:
                    # 获取子列的表头信息
                    child_header = self._get_child_header_with_worksheet(df, header_rows, col, parent_text, ws, start_row)
                    header_mapping[col] = child_header
        
        # 处理未合并的列
        for col in range(df.shape[1]):
            if col not in header_mapping:
                header = self._get_single_column_header_with_worksheet(df, header_rows, col, ws, start_row)
                header_mapping[col] = header
        
        return header_mapping

    def _get_child_header_with_worksheet(self, df: pd.DataFrame, header_rows: int, col: int, parent_text: str, ws: Worksheet, start_row: int) -> str:
        """获取合并单元格子列的表头。"""
        # 获取该列在表头行的所有非空值
        header_parts = []
        for row in range(header_rows):
            cell_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
            if cell_value.strip():
                header_parts.append(cell_value)
        
        # 构建层次结构表头
        if header_parts:
            if parent_text not in header_parts:
                header_parts.insert(0, parent_text)
            return "/".join(header_parts)
        else:
            return parent_text

    def _get_single_column_header_with_worksheet(self, df: pd.DataFrame, header_rows: int, col: int, ws: Worksheet, start_row: int) -> str:
        """获取单列的表头。"""
        header_parts = []
        for row in range(header_rows):
            cell_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
            if cell_value.strip():
                header_parts.append(cell_value)
        
        return "/".join(header_parts) if header_parts else ""

    def _df_to_html_multiheader(self, df: pd.DataFrame, header_rows: int, ws: Worksheet = None, start_row: int = 0) -> Tuple[str, List[str], List[Dict]]:
        """将多级表头的DataFrame转为HTML表格，正确处理合并单元格。"""
        if df.shape[0] == 0 or header_rows == 0:
            return "<table></table>", [], []
        
        # 获取合并单元格信息（从Excel文件获取）
        merged_cells = self._get_excel_merged_cells_with_worksheet(ws, start_row, header_rows) if ws else []

        # 构建表头映射
        header_mapping = self._build_header_mapping_with_worksheet(df, header_rows, merged_cells, ws, start_row) if ws else {}
        
        # 生成最终表头
        headers = []
        for col in range(df.shape[1]):
            if ws and col in header_mapping:
                header = header_mapping[col]
            else:
                # 回退到简单方法
                header_parts = []
                for row in range(header_rows):
                    cell_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                    if cell_value.strip():
                        header_parts.append(cell_value)
                header = "/".join(header_parts) if header_parts else ""
            headers.append(header)
        
        html = ["<table border='1'>"]
        
        # 构建合并单元格映射，用于HTML生成
        merge_map = {}
        for merge_info in merged_cells:
            start_col = merge_info["col"]
            colspan = merge_info["colspan"]
            for col_offset in range(colspan):
                col = start_col + col_offset
                if col < df.shape[1]:
                    merge_map[col] = {
                        "is_merged": col_offset == 0,  # 只有第一个单元格显示内容
                        "colspan": colspan if col_offset == 0 else 0,  # 只有第一个单元格有colspan
                        "text": merge_info["text"] if col_offset == 0 else ""
                    }
        
        # 多级表头 - 正确处理合并单元格
        for h in range(header_rows):
            html.append("<tr>")
            for col in range(df.shape[1]):
                if col in merge_map and merge_map[col]["is_merged"]:
                    # 合并单元格
                    colspan = merge_map[col]["colspan"]
                    text = merge_map[col]["text"]
                    html.append(f'<th colspan="{colspan}">{text}</th>')
                elif col in merge_map and not merge_map[col]["is_merged"]:
                    # 被合并的单元格，跳过
                    continue
                else:
                    # 普通单元格
                    cell_value = str(df.iloc[h, col]) if pd.notna(df.iloc[h, col]) else "-"
                    html.append(f"<th>{cell_value}</th>")
            html.append("</tr>")
        
        # 表体
        for i in range(header_rows, df.shape[0]):
            html.append("<tr>")
            for col in range(df.shape[1]):
                cell_value = str(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else "-"
                html.append(f"<td>{cell_value}</td>")
            html.append("</tr>")
        
        html.append("</table>")
        return "\n".join(html), headers, merged_cells

    def _df_to_markdown_multiheader(self, df: pd.DataFrame, header_rows: int, ws: Worksheet = None, start_row: int = 0) -> Tuple[str, List[str], List[Dict]]:
        """将多级表头的DataFrame转为Markdown表格，正确处理合并单元格。"""
        if df.shape[0] == 0 or header_rows == 0:
            return "", [], []
        
        # 获取合并单元格信息（从Excel文件获取）
        merged_cells = self._get_excel_merged_cells_with_worksheet(ws, start_row, header_rows) if ws else []

        # 构建表头映射
        header_mapping = self._build_header_mapping_with_worksheet(df, header_rows, merged_cells, ws, start_row) if ws else {}
        
        # 生成最终表头
        headers = []
        for col in range(df.shape[1]):
            if ws and col in header_mapping:
                header = header_mapping[col]
            else:
                # 回退到简单方法
                header_parts = []
                for row in range(header_rows):
                    cell_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                    if cell_value.strip():
                        header_parts.append(cell_value)
                header = "/".join(header_parts) if header_parts else ""
            headers.append(header)
        
        markdown_lines = []
        
        # 表头 - 使用处理后的表头，体现合并单元格的层次结构
        header_row = "| " + " | ".join(headers) + " |"
        markdown_lines.append(header_row)
        
        # 分隔线 - 使用标准格式并添加对齐方式
        separator_cells = []
        for col in range(df.shape[1]):
            # 根据数据类型确定对齐方式
            alignment = self._get_column_alignment(df, col, header_rows)
            separator_cells.append(alignment)
        separator_row = "| " + " | ".join(separator_cells) + " |"
        markdown_lines.append(separator_row)
        
        # 表体 - 过滤空行，确保数据行连续
        for i in range(header_rows, df.shape[0]):
            # 检查当前行是否为空行
            row_data = df.iloc[i]
            if row_data.notna().any():  # 只有当行中有非空数据时才添加
                data_row = "| " + " | ".join([str(x) if pd.notna(x) else "-" for x in row_data]) + " |"
                markdown_lines.append(data_row)
        
        return "\n".join(markdown_lines), headers, merged_cells

    def _get_column_alignment(self, df: pd.DataFrame, col: int, header_rows: int) -> str:
        """根据列的数据类型确定对齐方式"""
        # 检查数据行中的数据类型
        data_rows = df.iloc[header_rows:]
        if data_rows.empty:
            return "---"  # 默认左对齐
        
        # 获取该列的数据
        column_data = data_rows.iloc[:, col]
        
        # 检查是否为数值类型
        numeric_count = 0
        total_count = 0
        
        for value in column_data:
            if pd.notna(value):
                total_count += 1
                try:
                    # 尝试转换为数值
                    float(str(value).replace(',', '').replace('%', ''))
                    numeric_count += 1
                except (ValueError, TypeError):
                    pass
        
        # 如果超过70%的数据是数值，则右对齐
        if total_count > 0 and numeric_count / total_count > 0.7:
            return "---:"  # 右对齐
        else:
            return "---"   # 左对齐

    def _row_to_html(self, row: pd.Series, headers: List[str]) -> str:
        """将单行转为HTML表格。"""
        html = "<table border='1'><tr>"
        for cell in row:
            html += f"<td>{str(cell) if pd.notna(cell) else '-'}</td>"
        html += "</tr></table>"
        return html

    def _row_to_markdown(self, row: pd.Series, headers: List[str]) -> str:
        """将单行转为Markdown表格。"""
        # 修复：过滤空值，确保生成的markdown格式正确
        cells = []
        for cell in row:
            if pd.notna(cell) and str(cell).strip():
                cells.append(str(cell))
            else:
                cells.append("-")
        markdown = "| " + " | ".join(cells) + " |"
        return markdown

    def _get_merged_cells(self, ws: Worksheet) -> List[Dict[str, int]]:
        """获取合并单元格信息。"""
        merged: List[Dict[str, int]] = []
        for mc in ws.merged_cells.ranges:
            merged.append(
                {
                    "coord": str(mc.coord),
                    "min_row": mc.min_row,
                    "max_row": mc.max_row,
                    "min_col": mc.min_col,
                    "max_col": mc.max_col,
                }
            )
        return merged

    def _get_parent_table_info(
        self, headers: List[str], merged_cells: List[Dict[str, int]]
    ) -> str:
        """生成父表格信息描述。"""
        return f"表头: {headers} 合并单元格: {merged_cells}"

    def _get_context_for_table(
        self, df: pd.DataFrame, start_row: int, end_row: int
    ) -> str:
        """获取表格的上下文内容（前后非空行内容）。"""
        prev_content = self._get_nonempty_row_content(df, start_row - 1)
        next_content = self._get_nonempty_row_content(df, end_row + 1)
        context = f"上一行：{prev_content or ''}。下一行：{next_content or ''}"
        return context

    def _get_context_for_row(
        self, df: pd.DataFrame, row_idx: int, header_rows: int
    ) -> str:
        """获取行的上下文内容（前后行内容，排除表头）。"""
        prev_content = (
            self._get_nonempty_row_content(df, row_idx - 1)
            if row_idx > header_rows
            else ""
        )
        next_content = (
            self._get_nonempty_row_content(df, row_idx + 1)
            if row_idx < df.shape[0] - 1
            else ""
        )
        context = f"上一行：{prev_content or ''}。下一行：{next_content or ''}"
        return context

    def _get_nonempty_row_content(
        self, df: pd.DataFrame, row_idx: int
    ) -> Optional[str]:
        """获取指定行的内容（非空则拼接为字符串）。"""
        if 0 <= row_idx < df.shape[0]:
            row = df.iloc[row_idx]
            if row.notna().any():
                return " | ".join([str(x) if pd.notna(x) else "" for x in row])
        return None

    def _is_merged_cell_header(self, text: str) -> bool:
        """判断是否为合并单元格的表头。"""
        # 基于常见表头模式判断
        merged_patterns = ["营业总收入", "净利润", "每股收益", "净资产", "现金流量"]
        return any(pattern in text for pattern in merged_patterns)



    @classmethod
    def can_handle(cls, ext: str) -> bool:
        return ext.lower() == "xlsx"


def build_prompt_for_chunk(chunk: dict, with_context: bool = True) -> str:
    """
    根据分块类型和元数据动态生成Prompt，支持有无上下文。
    """
    chunk_type = chunk.get("type", "table_full")
    if with_context:
        PROMPTS = STRUCTURED_PROMPTS_WITH_CONTEXT
    else:
        PROMPTS = STRUCTURED_PROMPTS
    metadata = chunk.get("metadata", {})
    prompt = PROMPTS.get(chunk_type, PROMPTS["table_full"]).format(
        content=chunk.get("content", ""),
        sheet=metadata.get("sheet", ""),
        header=metadata.get("header", ""),
        parent_table_info=metadata.get("parent_table_info", ""),
        context=chunk.get("context", ""),
        row=metadata.get("row", ""),
        table_title=metadata.get("table_title", ""),
        paragraph_index=metadata.get("paragraph_index", ""),
    )
    return prompt


def get_system_prompt_for_chunk(chunk: dict) -> str:
    """
    根据分块类型获取系统提示词。
    """
    chunk_type = chunk.get("type", "table_full")
    return SYSTEM_PROMPTS.get(chunk_type, SYSTEM_PROMPTS["table_full"])


async def enhance_chunk(chunk: dict) -> dict:
    """
    调用LLM为分块生成description和keywords。
    """
    prompt = build_prompt_for_chunk(chunk)
    system_prompt = get_system_prompt_for_chunk(chunk)
    response = await zhipu_complete_async(
        prompt=prompt,
        api_key=LLM_CONFIG["api_key"],
        model=LLM_CONFIG["model"],
        temperature=LLM_CONFIG["temperature"],
        timeout=LLM_CONFIG["timeout"],
        max_tokens=LLM_CONFIG["max_tokens"],
        system_prompt=system_prompt,
    )
    result = parse_json_response(response)
    chunk.setdefault("metadata", {})["description"] = result.get("description", "")
    chunk["metadata"]["keywords"] = result.get("keywords", [])
    return chunk


async def enhance_all_chunks(chunks: List[Dict]) -> List[Dict]:
    """
    批量异步增强所有分块。
    """
    tasks = [enhance_chunk(chunk) for chunk in chunks]
    return await asyncio.gather(*tasks)
